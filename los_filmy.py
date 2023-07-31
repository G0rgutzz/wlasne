import random
from openpyxl import load_workbook
wb = load_workbook('puzzle.xlsx')
sheet = wb.active

nazwy = []
k = int(input("podaj ilość filmów: "))
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=k, max_col=1):
    for cell in row:
        nazwy.append(cell.value)

'''for i in range(k):
    a = input("podaj nazwę %s filmu: " % (i+1))
    nazwy.append(a)'''
wygrany = random.randint(1, k)
sheet.cell(row=3, column=3).value = wygrany

#  print("idziecie na film: %s" % nazwy[wygrany-1])
wb.save('puzzle.xlsx')

exit()
