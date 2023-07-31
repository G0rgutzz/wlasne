import calendar
import numpy as np
import pandas as pd
import seaborn as sn
from datetime import datetime
import matplotlib.pyplot as plt
import scipy.stats as stats
from openpyxl import Workbook
wb = Workbook()
sheet = wb.active

bs = pd.read_csv('bikesharing_prepared.csv')  # index_col=0, sep='', header=0, decimal='.')
# print(bs.info)

# print(bs.shape)

# typy danych
# print(bs.dtypes)
# print(pd.DataFrame(bs.dtypes.value_counts()))

# Kilka pierwszych obserwacji
# print(bs.head())

# wybranie kilku kolumn
# print(bs[['count', 'season', 'temp']])
# print(bs.columns[0], '\t', bs.columns[7])  # nazwy kolumn
# print(bs[bs.columns[7]])  # wywołanie po numerze kolumny, nie podaje nazwy kolumny
# print(bs.columns[2], bs.columns[5], bs.columns[7], bs.columns[10])
bs['season2'] = 1
bs['przyklad'] = bs['temp']*bs['humidity']
print(bs.head())
i = 0
for row in sheet.iter_rows(min_row=2, min_col=3, max_row=40, max_col=3):
    for cell in row:
        sheet.cell(row=i+1, column=1).value = bs
        i = i+1
wb.save("dane.xlsx")
# nazwy kolumn
# print(bs.columns)

# kasowanie kolumn
'''bs.drop(columns = (['nazwy kolumn'], inplace = True) albo w sposób opisany niżej
bs = bs.drop(columns = (['nazwy kolumn'], inplace = True)'''

''' loc - indeks etykietowy, wiersze true pokaże, false zignoruje
query to nakładka na loc w formie stringa, żeby łatwiej zapisywać polecenia
'''
'''
print(bs.loc[(bs.season == 'Spring') & (bs.temp > 10)].head())
print(bs.query('season == "Spring" and temp > 10').head())'''

# grupowanie
# bsGR_hour_seas = (bs.groupby(["hour", "season"]))
# bsGR_hour_wday = bs.groupby(["hour", "weekday"])

# print(bsGR_hour_seas)
# print(bsGR_hour_wday)
