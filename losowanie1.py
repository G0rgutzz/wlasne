import random
from openpyxl import load_workbook
wb = load_workbook('uczniowie.xlsx')
sheet = wb.active

'''sheet.cell(row=1, column=1).value = "numer"
sheet.cell(row=1, column=2).value = "nazwisko"
sheet.cell(row=1, column=3).value = "wylosowane"'''
nazwisko = []
numer = []
liczby = []
j = 0
i = 0

k = int(input("Ilu uczniów jest w klasie: "))
for row in sheet.iter_rows(min_row=2, min_col=3, max_row=40, max_col=3):
    for cell in row:
        cell.value = None

for row in sheet.iter_rows(min_row=2, min_col=2, max_row=k+1, max_col=2):
    for cell in row:
        nazwisko.append(cell.value)
        numer.append(i+1)
        i = i+1

'''for i in range(k):
    imie = (input("Podaj imię i nazwisko %s ucznia: " % (i+1)))
    sheet.cell(row=i+2, column=2).value = imie
    sheet.cell(row=i+2, column=1).value = i+1
    nazwisko.append(imie)
    numer.append(i+1)
    i = i+1'''
while len(liczby) < len(numer):
    liczba = random.randint(1, k)
    if liczba != j+1 and liczba not in liczby:
        liczby.append(liczba)
        los = nazwisko[liczba-1]
        sheet.cell(row=j+2, column=3).value = los
        j = j+1
    else:
        liczba = random.randint(1, k)

# print(nazwisko, "\n", numer, "\n", liczby)
wb.save('uczniowie.xlsx')

exit()
